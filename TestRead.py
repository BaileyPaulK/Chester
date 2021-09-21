from openpyxl import load_workbook
workbook = load_workbook(filename = "Data.xlsx")
sheet = workbook.active

print(workbook.sheetnames)

for x in workbook.sheetnames:
    print(x)

while True:
    cell = input("Please Input Cell: ")
    if cell == "exit":
        break
    print(sheet[cell].value)
