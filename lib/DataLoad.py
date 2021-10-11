from openpyxl import load_workbook
from lib import Board
from lib import Part
from lib import Input

def Exists (id, list):
    for item in list:
        if item.id == id:
            return item
    return False

def DataLoad (file, boards, parts):
    wb = load_workbook(filename = "data/" + file)     #loads in workbook
    sheet = wb.active                       #sets first sheet or tab as sheet
    row = 2
    while sheet['A' + str(row)].value != None:
        boardid     = sheet['C' + str(row)].value
        #print(boardid)
        partid      = sheet['F' + str(row)].value
        #print(partid)
        partqty     = sheet['E' + str(row)].value
        #print(partqty)
        partstock   = sheet['I' + str(row)].value
        #print(partstock)
        board = Exists(boardid, boards)
        if board == False:
            board = Board.Board(boardid)
            boards.append(board)
        part = Exists(partid, parts)
        if part == False:
            part = Part.Part(partid, partstock)
            parts.append(part)
        board.AddPart(part, partqty)
        row += 1 

def InputLoad (inputs, boards):
    wb = load_workbook(filename = "input.xlsx")
    sheet = wb.active
    row = 2
    print("loading input")
    while sheet['A' + str(row)].value != None:
        boardName   = sheet['A' + str(row)].value
        qty         = sheet['B' + str(row)].value
        #print(row) #i was dumb and forgot row += 1, took this for me to figure it out, and it stays to remind me of the shame :,)
        inputs.append(Input.Input(boardName, qty, row, boards))
        row += 1
    return wb