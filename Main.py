import os
from lib import DataLoad
from lib import Board
from lib import Part
from lib import Input

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
clearFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFFFFFF',
                   fill_type='solid')
no_fill = PatternFill(fill_type=None)
boards = []
parts = []
inputs = []

for file in os.listdir("data"):     #load data from all files in dir 
    if (file.endswith(".xlsx") and not file.startswith("~")):
        print("Loading From: " + file)
        DataLoad.DataLoad(file, boards, parts)
wb = DataLoad.InputLoad(inputs, boards)  #Load input
sheet = wb.active

for order in inputs:        #calculate stock v qty
    #print("Pulling " + str(order.board.id))
    order.Pull()

for part in parts:
    print("partid: " + str(part.id))
    stock   = part.stock
    qty     = part.qty
    print("stock: " + str(stock) + ", qty: " + str(qty))
    part.IsOut()

#output (this may get messy)
for order in inputs:
    if order.IsOut(): 
        sheet.cell(row = order.row, column = 2).fill = redFill
        #sheet.cell(row = order.row, column = 3, value = "FLAG")
        print("Flag: row= " + str(order.row) + ", id= " + str(order.board.id))
    else:
        sheet.cell(row = order.row, column = 2).fill = no_fill
        #sheet.cell(row = order.row, column = 3).value = None

#delete all none input sheets
for sheetName in wb.get_sheet_names():
    if not (sheet == 'Input'):
        wb.remove_sheet(wb.get_sheet_by_name(sheet))

for part in parts:
    if part.flag:
        part.sheet = wb.create_sheet("Part: " + str(part.id))
        part.sheet.append(["PartID", part.id, "Stock", part.stock, "Ordered", part.qty])
        part.sheet.append(["BoardID", "Board Qty", "Parts needed"])

for order in inputs:
    if order.flag:
        order.sheet = wb.create_sheet("Board: " + str(order.board.id))
        order.sheet.append(["BoardID", order.board.id])
        order.sheet.append(["Parts Causing Issues"])
        order.sheet.append(["PartID", "Needed for order", "Stock"])
        for part in order.board.parts:
            if part[0].flag:
                qty = order.qty * part[1]
                order.sheet.append([part[0].id, qty, part[0].stock])
                part[0].sheet.append([order.board.id, order.qty, qty])
wb.save('input.xlsx')